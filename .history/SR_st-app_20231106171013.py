import streamlit as st
from openpyxl import load_workbook
import pandas as pd
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

st.set_page_config(
    page_title="Slide-Rule",
    page_icon="📏",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.extremelycoolapp.com/help',
        'Report a bug': "https://www.extremelycoolapp.com/bug",
        'About': "# This is a header. This is an *extremely* cool app!"
    }
)

st.title("Slide Rule")

st.write(
    """My first Streamlit app
    """
)

st.header("Coucou")

data_file = st.sidebar.file_uploader("Upload Excel file", ["xls", "xlsx"])  
if data_file:
    file_details = {
        "Filename":data_file.name,
        "FileType":data_file.type,
        "FileSize":data_file.size
        }

    wb = load_workbook(data_file)

    ## Show Excel file
    st.sidebar.subheader("File details:")
    st.sidebar.json(file_details,expanded=False)
    st.sidebar.markdown("----")



# myfile = st.file_uploader("Upload Excel file", type=["xls", "xlsx"])

# if myfile:
#     st.info(f"File uploaded: {myfile.name}")
#     #st.info(f"Sheet names: {myfile.sheetnames}")
#     #my_df = pd.read_excel(myfile, header = 0)
#     #st.dataframe(filter_dataframe(my_df))

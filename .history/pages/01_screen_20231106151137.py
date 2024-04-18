import pandas as pd
import streamlit as st

from openpyxl import load_workbook
from io import BytesIO


from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)


st.header("Sheet view")
data_file = st.sidebar.file_uploader("Upload Excel file",type=['xlsx'])  

if data_file:
    file_details = {
        "Filename":data_file.name,
        "FileType":data_file.type,
        "FileSize":data_file.size}

    wb = load_workbook(data_file)

    ## Show Excel file
    st.sidebar.subheader("File details:")
    st.sidebar.json(file_details,expanded=False)
    st.sidebar.markdown("----")

    ## Select sheet
    sheet_selector = st.sidebar.selectbox("Selected sheet:", wb.sheetnames)     
    df = pd.read_excel(data_file,sheet_selector)
    st.markdown(f"### Currently Selected: `{sheet_selector}`")
    st.write(df)

    ## Do something after a button
    doLogic_btn = st.button("➕")
    if doLogic_btn:
        df2 = df.sum().transpose()
        st.write(df2)

        # Do something more after the previous button
        # >> But this will fail because the button will go back to _False_ 
        # >> so nothing will be shown afterwards
        another_btn = st.checkbox("Another +")
        if another_btn:
            df3 = df2.sum()
            st.write(df3)
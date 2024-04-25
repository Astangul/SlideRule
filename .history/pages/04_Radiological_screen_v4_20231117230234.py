import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

st.set_page_config(
    page_title="Slide-Rule",
    page_icon="📏",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.extremelycoolapp.com/help',
        'Report a bug': "https://www.extremelycoolapp.com/bug",
        'About': "# This is a header. This is an *extremely* cool app!"
    }
)

def filter_dataframe_multiselect(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    # modify = st.checkbox("Add filters")

    # if not modify:
    #     return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns, default=list(df.columns))
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.multiselect(
                    f"Values for {column}",
                    df[column].unique(),
                    #default=list(df[column].unique()),
                    max_selections = 1
                )
                df = df[df[column].isin(user_cat_input)]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].str.contains(user_text_input)]

    return df

def filter_dataframe_selectbox(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    # modify = st.checkbox("Add filters")

    # if not modify:
    #     return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns, default=list(df.columns))
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.selectbox(
                    label=f"Values for {column}",
                    options=df[column].unique().tolist(),
                    index=0
                    # f"Values for {column}",
                    # df[column].unique(),
                    # #default=list(df[column].unique()),
                    # max_selections = 1
                )
                
                df = df[df[column].isin([user_cat_input])]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].str.contains(user_text_input)]

    return df

def scatter_chart(x_data, y_data, err_y_pc):
    fig = go.Figure(data=go.Scatter(
        x=x_data,
        y=y_data,
        error_y=dict(
            type = 'data', # value of error bar given in data coordinates
            array = err_y_pc*y_data,
            visible = True)
        ))
    fig.update_xaxes(type="log", showgrid=True)
    fig.update_xaxes(minor = dict(ticks="inside", ticklen=6, griddash='dot', showgrid=True))
    fig.update_yaxes(type='log', showgrid=True, tickformat='.2e')
    st.plotly_chart(fig, theme="streamlit")


st.title("Slide Rule")

st.write(
    """My first Streamlit app
    """
)

data_file = "./DB/All-at-once_DB.xlsx"

# @st.cache_data
if data_file:
    wb = load_workbook(data_file)
    ## Select sheet
    sheet_selector = st.sidebar.selectbox("Selected sheet:", wb.sheetnames)     
    df_SR_original = pd.read_excel(data_file, sheet_selector, header = 0)


#selected_options = st.multiselect("Choose screen", df_SR_original['Screen'].unique(), max_selections = 1)
#selected_options = st.selectbox("Choose screen", df_SR_original['Screen'].unique())





# https://arnaudmiribel.github.io/streamlit-extras/extras/grid/
# https://arnaudmiribel.github.io/streamlit-extras/extras/chart_container/
# https://lukasmasuch-streamlit-pydantic-playgroundplayground-app-711bhu.streamlit.app/


#Mettre le tableau filtre / graphe dans des onglets

with st.expander("See explanation"):
    st.write(
        """The chart above shows some numbers I picked for you.
        I rolled actual dice for these, so they're *guaranteed* to
        be random."""
    )
    st.image("https://static.streamlit.io/examples/dice.jpg")



# with st.expander("Select reference case"):
#     st.selectbox("Choose screen", df_SR_original['Screen'].unique())


# with st.expander("Select reference case"):
#     st.dataframe(filter_dataframe_test(df_SR_original))

# tabs = st.tabs(['Filtered Data', 'Filter Data'])

# with tabs[0]:  # First tab for filtered data
#     filtered_data = filter_dataframe_test(df_SR_original)
#     st.write(filtered_data)  # Display filtered dataframe in the first tab

# with tabs[1]:  # Second tab for filtering options
#     filter_dataframe_test(df_SR_original)  # Display filter options in the second tab

# tabs = st.tabs(['Filtered Data', 'Filter Data'])
# with st.expander("Chosse reference case"):
#     with tabs[0]:
#         filter_dataframe(df_SR_original)

tab1, tab2, tab3 = st.tabs(["Reference Case", "Data", "Graph"])
with tab1:
    df_reference = filter_dataframe_selectbox(df_SR_original)
with tab2:
    st.dataframe(df_reference)
with tab3:
    scatter_chart(df_reference['Distance (m)'], df_reference['Dose (Gy)'], 2*df_reference['1s uncertainty'])


# distance = df_reference['Distance (m)']
# dose = df_reference['Dose (Gy)']
# two_sig = 2*(df_reference['1s uncertainty']*df_reference['Dose (Gy)'])
# scatter_chart(distance, dose, two_sig)



t1, t2 = st.tabs(["titi", "toto"])


with st.expander("Expander"):
    with t1:
        st.write("Content of Tab 1")
    with t2:
        st.write("Content of Tab 2")




with st.expander("Expander"):
    with st.tabs("Tabs"):
        with st.tab("Tab 1"):
            st.write("Content of Tab 1")
        with st.tab("Tab 2"):
            st.write("Content of Tab 2")
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook


st.set_page_config(
    page_title="Slide-Rule",
    page_icon="📏",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.extremelycoolapp.com/help',
        'Report a bug': "https://www.extremelycoolapp.com/bug",
        'About': "# This is a header. This is an *extremely* cool app!"
    }
)


import sys
from pathlib import Path

# Ajoutez le répertoire parent à sys.path pour rendre func_utils importable
sys.path.append(str(Path(__file__).resolve().parent.parent))

from utils.st_filter_dataframe import df_multiselect_filters_v2


def scatter_chart(x_data, y_data, err_y_pc):
    fig = go.Figure(data=go.Scatter(
        x=x_data,
        y=y_data,
        error_y=dict(
            type = 'data', # value of error bar given in data coordinates
            array = err_y_pc*y_data,
            visible = True)
        ))
    fig.update_xaxes(type="log", showgrid=True, title="Distance (m)")
    fig.update_xaxes(minor = dict(ticks="inside", ticklen=6, griddash='dot', showgrid=True))
    fig.update_yaxes(type='log', showgrid=True, tickformat='.2e', title="Dose (Gy)")

    # Ajuste la taille du graphe (hauteur et largeur)
    fig.update_layout(height=600, width=900)  # Modifie ces valeurs selon tes besoins

    st.plotly_chart(fig, theme="streamlit")

st.title("Slide Rule")

st.write(
    """My first Streamlit app
    """
)

data_file = "./DB/All-at-once_DB.xlsx"

# @st.cache_data
if data_file:
    wb = load_workbook(data_file)
    ## Select sheet
    sheet_selector = st.sidebar.selectbox("Selected sheet:", wb.sheetnames)     
    df_SR_original = pd.read_excel(data_file, sheet_selector, header = 0)

# 1/ _____ REFERENCE CASE _____
with st.expander("1/ Select Reference Case"):
    tab1, tab2, tab3 = st.tabs(["📑 Filters", "🔢 Data", "📈 Graph"])
    
    with tab1:
        # Contenu pour l'onglet "Reference Case"
        SR_reference_df, SR_reference_filters = df_multiselect_filters_v2(df_SR_original, default_columns=['Fissile', 'Case'])
    
    with tab2:
        # Contenu pour l'onglet "Data"
        st.dataframe(SR_reference_df)
    
    with tab3:
        # Contenu pour l'onglet "Graph"
        scatter_chart(SR_reference_df['Distance (m)'], SR_reference_df['Dose (Gy)'], 2*SR_reference_df['1s uncertainty'])

st.write("Reference case selected filters :", SR_reference_filters)

# 2/ _____ TEST CASE _____
#toto_df, SR_test_filters = df_multiselect_filters_v2(df_SR_original)
# Affichage des filtres sélectionnés
st.write("Filtres sélectionnés :", SR_test_filters)
st.write("DataFrame filtré :", toto_df)



filtered_df = df_SR_original.copy()  # Copie du DataFrame d'origine pour commencer avec toutes les données

for column, values in SR_reference_filters.items():
    if column in SR_test_filters:
        test_values = SR_test_filters[column]
        if isinstance(test_values, list):  # Vérifie si les valeurs sont stockées dans une liste
            filtered_df = filtered_df[filtered_df[column].isin(test_values)]
        elif isinstance(test_values, tuple) and len(test_values) == 2:  # Pour les plages (par exemple, dates)
            filtered_df = filtered_df[filtered_df[column].between(test_values[0], test_values[1])]
    else:
        if isinstance(values, list) or isinstance(values, tuple):  # Vérifie si les valeurs sont stockées dans une liste ou un tuple
            filtered_df = filtered_df[filtered_df[column].isin(values)]
        else:
            filtered_df = filtered_df[filtered_df[column] == values]

# filtered_df contient maintenant les données filtrées selon les spécifications
st.write("DataFrame filtré :", filtered_df)

# ___________________________________________________________________________________
# https://arnaudmiribel.github.io/streamlit-extras/extras/grid/
# https://arnaudmiribel.github.io/streamlit-extras/extras/chart_container/
# https://lukasmasuch-streamlit-pydantic-playgroundplayground-app-711bhu.streamlit.app/


# updated_filters = SR_reference_filters.copy()  # On commence avec les filtres de SR_reference_filters

# for column, values in SR_test_filters.items():
#     if column in updated_filters:
#         updated_filters[column] = values  # Met à jour les valeurs pour les clés existantes dans SR_reference_filters

# filtered_df = df_SR_original.copy()  # Copie du DataFrame d'origine pour commencer avec toutes les données

# for column, values in updated_filters.items():
#     if isinstance(values, list):  # Vérifie si les valeurs sont stockées dans une liste
#         filtered_df = filtered_df[filtered_df[column].isin(values)]
#     elif isinstance(values, tuple) and len(values) == 2:  # Pour les plages (par exemple, dates)
#         filtered_df = filtered_df[filtered_df[column].between(values[0], values[1])]
#     else:
#         filtered_df = filtered_df[filtered_df[column] == values]

# # filtered_df contient maintenant les données filtrées selon les spécifications
# st.write("DataFrame filtré :", filtered_df)


# Todo list
# Renvoyer le tableau pour lesquels les convergences de calcul sont trop élevées. 
# Renvoyer les cas pour lesquels le ratios est supérieur à 1+epsilon





import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

st.set_page_config(
    page_title="Slide-Rule",
    page_icon="📏",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.extremelycoolapp.com/help',
        'Report a bug': "https://www.extremelycoolapp.com/bug",
        'About': "# This is a header. This is an *extremely* cool app!"
    }
)


def dataframe_selectbox(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    # modify = st.checkbox("Add filters")

    # if not modify:
    #     return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns, default=list(df.columns))
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.selectbox(
                    label=f"Values for {column}",
                    options=df[column].unique().tolist(),
                    index=0
                    # f"Values for {column}",
                    # df[column].unique(),
                    # #default=list(df[column].unique()),
                    # max_selections = 1
                )
                
                df = df[df[column].isin([user_cat_input])]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].str.contains(user_text_input)]

    return df


def dataframe_multiselect(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    # modify = st.checkbox("Add filters")

    # if not modify:
    #     return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns, max_selections = 1)
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.multiselect(
                    f"Values for {column}",
                    df[column].unique(),
                    default=list(df[column].unique()),
                    #max_selections = 1
                )
                df = df[df[column].isin(user_cat_input)]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].str.contains(user_text_input)]

    return df



def dataframe_selectbox_filters(df: pd.DataFrame) -> (pd.DataFrame, dict):
    # """
    # Adds a UI on top of a dataframe to let viewers filter columns

    # Args:
    #     df (pd.DataFrame): Original dataframe

    # Returns:
    #     pd.DataFrame: Filtered dataframe
    # """
    # # modify = st.checkbox("Add filters")

    # # if not modify:
    # #     return df

    # df = df.copy()

    # # Try to convert datetimes into a standard format (datetime, no timezone)
    # for col in df.columns:
    #     if is_object_dtype(df[col]):
    #         try:
    #             df[col] = pd.to_datetime(df[col])
    #         except Exception:
    #             pass

    #     if is_datetime64_any_dtype(df[col]):
    #         df[col] = df[col].dt.tz_localize(None)
    
    
    filters = {}  # Variable pour stocker les filtres sélectionnés

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns, default=list(df.columns))
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")

            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.selectbox(
                    label=f"Values for {column}",
                    options=df[column].unique().tolist(),
                    index=0
                )
                filters[column] = user_cat_input  # Stocke la valeur du filtre pour cette colonne
                df = df[df[column].isin([user_cat_input])]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                )
                filters[column] = user_num_input  # Stocke la valeur du filtre pour cette colonne
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    filters[column] = user_date_input  # Stocke la valeur du filtre pour cette colonne
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    filters[column] = user_text_input  # Stocke la valeur du filtre pour cette colonne
                    df = df[df[column].str.contains(user_text_input)]

    return df, filters  # Retourne à la fois le DataFrame filtré et le dictionnaire de filtres


def dataframe_multiselect_filters(df: pd.DataFrame) -> (pd.DataFrame, dict):
    # """
    # Adds a UI on top of a dataframe to let viewers filter columns

    # Args:
    #     df (pd.DataFrame): Original dataframe

    # Returns:
    #     pd.DataFrame: Filtered dataframe
    # """
    # # modify = st.checkbox("Add filters")

    # # if not modify:
    # #     return df

    # df = df.copy()

    # # Try to convert datetimes into a standard format (datetime, no timezone)
    # for col in df.columns:
    #     if is_object_dtype(df[col]):
    #         try:
    #             df[col] = pd.to_datetime(df[col])
    #         except Exception:
    #             pass

    #     if is_datetime64_any_dtype(df[col]):
    #         df[col] = df[col].dt.tz_localize(None)
    
    
    filters = {}  # Variable pour stocker les filtres sélectionnés

    modification_container = st.container()

    with modification_container:
        #to_filter_columns = st.multiselect("Filter dataframe on", df.columns, default=list(df.columns))
        to_filter_column = st.selectbox("Filter dataframe on", df.columns, index=0)
        to_filter_columns = [to_filter_column] if to_filter_column else list(df.columns)

        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.multiselect(
                    f"Values for {column}",
                    df[column].unique(),
                    default=list(df[column].unique()),
                    #max_selections = 1
                )
                filters[column] = user_cat_input  # Stocke la valeur du filtre pour cette colonne
                df = df[df[column].isin(user_cat_input)]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                )
                filters[column] = user_num_input  # Stocke la valeur du filtre pour cette colonne
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    filters[column] = user_date_input  # Stocke la valeur du filtre pour cette colonne
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    filters[column] = user_text_input  # Stocke la valeur du filtre pour cette colonne
                    df = df[df[column].str.contains(user_text_input)]

    return df, filters  # Retourne à la fois le DataFrame filtré et le dictionnaire de filtres

def scatter_chart(x_data, y_data, err_y_pc):
    fig = go.Figure(data=go.Scatter(
        x=x_data,
        y=y_data,
        error_y=dict(
            type = 'data', # value of error bar given in data coordinates
            array = err_y_pc*y_data,
            visible = True)
        ))
    fig.update_xaxes(type="log", showgrid=True, title="Distance (m)")
    fig.update_xaxes(minor = dict(ticks="inside", ticklen=6, griddash='dot', showgrid=True))
    fig.update_yaxes(type='log', showgrid=True, tickformat='.2e', title="Dose (Gy)")

    # Ajuste la taille du graphe (hauteur et largeur)
    fig.update_layout(height=600, width=900)  # Modifie ces valeurs selon tes besoins

    st.plotly_chart(fig, theme="streamlit")


st.title("Slide Rule")

st.write(
    """My first Streamlit app
    """
)

data_file = "./DB/All-at-once_DB.xlsx"

# @st.cache_data
if data_file:
    wb = load_workbook(data_file)
    ## Select sheet
    sheet_selector = st.sidebar.selectbox("Selected sheet:", wb.sheetnames)     
    df_SR_original = pd.read_excel(data_file, sheet_selector, header = 0)


with st.expander("Select Reference Case"):
    tab1, tab2, tab3 = st.tabs(["Filter", "Data", "Graph"])
    
    with tab1:
        # Contenu pour l'onglet "Reference Case"
        df_reference = dataframe_selectbox(df_SR_original)
    
    with tab2:
        # Contenu pour l'onglet "Data"
        st.dataframe(df_reference)
    
    with tab3:
        # Contenu pour l'onglet "Graph"
        scatter_chart(df_reference['Distance (m)'], df_reference['Dose (Gy)'], 2*df_reference['1s uncertainty'])




# Mettons en place un DataFrame factice pour l'exemple
data = {
    'Name': ['Alice', 'Bob', 'Charlie'],
    'Age': [25, 30, 35],
    'Gender': ['Female', 'Male', 'Male'],
    'Salary': [50000, 60000, 70000],
    'Join_date': ['2020-01-01', '2015-05-15', '2010-10-10']
}

df = pd.DataFrame(data)


#filtered_df, filters_selected = dataframe_selectbox_filters(df)
filtered_df, filters_selected = dataframe_multiselect_filters(df)

st.write("DataFrame non filtré:", df)



# # Affichage du DataFrame  filtré
# st.write("DataFrame filtré:", filtered_df)

# Affichage des filtres sélectionnés
st.write("Filtres sélectionnés :", filters_selected)

# Exemple d'utilisation des filtres sélectionnés
if 'Age' in filters_selected:
    st.write("Filtre sur l'âge :", filters_selected['Age'])

# Application des filtres sur le DataFrame initial
for col, val in filters_selected.items():
    filtered_df = filtered_df[filtered_df[col] == val]

# Affichage du DataFrame filtré
st.write("DataFrame filtré :", filtered_df)


# https://arnaudmiribel.github.io/streamlit-extras/extras/grid/
# https://arnaudmiribel.github.io/streamlit-extras/extras/chart_container/
# https://lukasmasuch-streamlit-pydantic-playgroundplayground-app-711bhu.streamlit.app/
